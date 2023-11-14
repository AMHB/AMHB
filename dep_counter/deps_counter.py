import pandas as pd
from openpyxl import load_workbook

# Load the Excel file
excel_file = 'site_list.xlsx'  # Replace with your file path
sheet_name = 'Sheet1'  # Replace with your sheet name

# Read the Excel file into a pandas DataFrame
df = pd.read_excel(excel_file, sheet_name=sheet_name)

# Group by 'sites_A' and aggregate the dependent values into a list
dependents_list = df.groupby('sites_A')['sites_B'].agg(list).reset_index()
dependents_list.columns = ['Sites', 'Dependents']

# Count the number of dependents for each value in column A
dependents_count = df['sites_A'].value_counts().reset_index()
dependents_count.columns = ['Sites', 'Number of Dependents']

# Add missing values from column B
missing_sites = set(df['sites_B']) - set(df['sites_A'])
missing_dependents = pd.DataFrame({
    'Sites': list(missing_sites),
    'Number of Dependents': 0,
    'Dependents': [''] * len(missing_sites)
})

# Concatenate DataFrames to combine counts and missing values
dependents_count = pd.concat([dependents_count, missing_dependents]).reset_index(drop=True)

# Merge count and dependent lists on 'Sites'
merged_df = pd.merge(dependents_count, dependents_list, on='Sites', how='left')

# Sort by 'Sites'
merged_df = merged_df.sort_values(by='Sites').reset_index(drop=True)

# Write the output to a new sheet in the same Excel file
with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
    merged_df.to_excel(writer, sheet_name='Dependents', index=False)

    # Writing dependents to separate columns starting from column E
    if 'Dependents' in merged_df.columns:
        wb = load_workbook(excel_file)
        sheet = wb['Dependents']
        for index, row in merged_df.iterrows():
            dependents = row['Dependents']
            if dependents:
                sorted_dependents = sorted(dependents)
                for i, dependent in enumerate(sorted_dependents):
                    sheet.cell(row=index + 2, column=i + 5, value=dependent)  # Column E is index 5
        wb.save(excel_file)
    