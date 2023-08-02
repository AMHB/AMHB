import pandas as pd
import numpy as np
from xlwt import *
import jinja2
from jinja2 import Template
from openpyxl import Workbook

wb = Workbook()
sheet = wb.active

R_rE = 22
R_rK = 42
freq = [13,18,23,26,28,38,80]
n = [1,2,3,4,5,6,7,8,9,10]
k= [0.03041, 0.07078, 0.1286, 0.1724, 0.2051, 0.4001, 1.1827]
a = [1.1586, 1.0818, 1.0214, 0.9884, 0.9679, 0.8816, 0.7096]


col1 = "Frequency"
col2 = "Distance"
col3 = "Rain_Fading_Zone_E(dBm)"
col4 = "Rain_Fading_Zone_K(dBm)"


distance = 0
F_rK = 0
F_rE = 0
cnt = 1



H1 = sheet.cell(row = 1, column = 1)
H1.value = " Link_Distance(km)"
H2 = sheet.cell(row= 1, column = 2)
H2.value = " Frequency(GHz)"
H3 = sheet.cell(row = 1, column = 3)
H3.value = "Rain_Attenuation_Zone_E(dBm)"
H4 = sheet.cell(row = 1, column = 4)
H4.value = "Rain_Attenuation_Zone_K(dBm)"
H4 = sheet.cell(row = 1, column = 5)
H4.value = "Planning_Status"



for distance in range (1,11):
    
    for i in range (7):
        cnt+=1
        template = Template('sheet{{c}}')
        x = template.render(c=i)
        F_rE = distance* k[i] * R_rE**a[i] 
        F_rK = distance* k[i] * R_rK**a[i]
        #x =  print ("The Rain Fading in E zone in the frequency of ",freq[i],'GHz ','and link distance of ', distance ,'Km is = ',F_rE, ' dBm')
        #y =  print ("The Rain Fading in E zone in the frequency of ",freq[i],'GHz ','and link distance of ', distance ,'Km is = ',F_rK, ' dBm')
        #----------------------------------------------
        ## with xlwt
        # sheet1 = wb.add_sheet('Sheet{{c}}')
        # #sheet1.write(row,col, data, style)
        # # sheet1.write(1, 0, 'Frequency')
        # # sheet1.write(2, 0, 'Frequency')
        # # sheet1.write(3, 0, 'Frequency')
        # # sheet1.write(4, 0, 'Frequency')
        # # wb.save('sample_data3.xls')
        #-----------------------------------------------
        ## with openpyxl
        #data = pd.DataFrame({col1:freq,col2:distance,col3:F_rE,col4:F_rK})
        #data.to_excel('test111.xlsx', sheet_name='sheet1')
      
        #-----------------------------------------------
        c1 = sheet.cell(row = cnt, column = 1)
        c1.value = distance
        c2 = sheet.cell(row= cnt, column = 2)
        c2.value = freq[i]
        c3 = sheet.cell(row = cnt, column = 3)
        c3.value = F_rE
        if F_rE >= 40:
            c5 = sheet.cell(row = cnt, column = 5)
            c5.value = "Impossible"
        elif 30 <F_rE < 40:
            c5 = sheet.cell(row = cnt, column = 5)
            c5.value = "Risky"
        else:
            c5 = sheet.cell(row = cnt, column = 5)
            c5.value = "Possible" 
        c4 = sheet.cell(row = cnt, column = 4)
        c4.value = F_rK
        if F_rK >= 40:
            c5 = sheet.cell(row = cnt, column = 5)
            c5.value = "Impossible"
        elif 30 <F_rK < 40:
            c5 = sheet.cell(row = cnt, column = 5)
            c5.value = "Risky"
        else:
            c5 = sheet.cell(row = cnt, column = 5)
            c5.value = "Possible"     
        wb.save("Rain_Attenation.xlsx")